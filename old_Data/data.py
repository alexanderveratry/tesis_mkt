import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Leer el archivo JSON
df = pd.read_json('response_1755622727310.json')

# Convertir la columna FECHA a tipo datetime y eliminar zona horaria
df['FECHA'] = pd.to_datetime(df['FECHA']).dt.tz_localize(None)

# Ordenar por INDICE, DIVIDENDO y FECHA
df_ordenado = df.sort_values(['INDICE', 'DIVIDENDO', 'FECHA'])

# Seleccionar columnas en el orden solicitado
columnas = [
    'INDICE', 'PERIODO', 'FECHA', 'DIVIDENDO', 'IND_ACT', 'IND_VAR',
    'IND_RENT_7DIAS', 'IND_RENT_30DIAS', 'IND_RENT_52SEM', 'IND_RENT_ANO_AC', 'NUM_ACC_COM'
]
df_final = df_ordenado[columnas]

# Crear una copia para CSV (con formato string para fecha)
df_csv = df_final.copy()
df_csv['FECHA'] = df_csv['FECHA'].dt.strftime('%Y-%m-%d %H:%M:%S')

# Guardar como CSV con cabecera, sin índice, usando punto y coma como separador y formato decimal correcto
df_csv.to_csv('resultado_limpio.csv', index=False, header=True, sep=';', float_format='%.8f')

# Crear archivo Excel con formato mejorado
excel_filename = 'datos_indices_financieros.xlsx'

# Guardar el DataFrame en Excel
with pd.ExcelWriter(excel_filename, engine='openpyxl', date_format='YYYY-MM-DD HH:MM:SS') as writer:
    df_final.to_excel(writer, sheet_name='Datos Índices', index=False)
    
    # Obtener el workbook y worksheet para aplicar formato
    workbook = writer.book
    worksheet = writer.sheets['Datos Índices']
    
    # Estilo para el encabezado
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Estilo para bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aplicar formato al encabezado
    for col_num, column_title in enumerate(columnas, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Aplicar formato a las celdas de datos
    for row in range(2, len(df_final) + 2):
        for col in range(1, len(columnas) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            
            # Alineación específica por tipo de columna
            if col in [1, 2, 4]:  # INDICE, PERIODO, DIVIDENDO
                cell.alignment = Alignment(horizontal='center')
            elif col == 3:  # FECHA
                cell.alignment = Alignment(horizontal='center')
            else:  # Columnas numéricas
                cell.alignment = Alignment(horizontal='right')
    
    # Ajustar ancho de columnas
    column_widths = {
        'A': 15,  # INDICE
        'B': 10,  # PERIODO
        'C': 20,  # FECHA
        'D': 12,  # DIVIDENDO
        'E': 15,  # IND_ACT
        'F': 12,  # IND_VAR
        'G': 18,  # IND_RENT_7DIAS
        'H': 18,  # IND_RENT_30DIAS
        'I': 18,  # IND_RENT_52SEM
        'J': 18,  # IND_RENT_ANO_AC
        'K': 15   # NUM_ACC_COM
    }
    
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

print('Archivos generados correctamente:')
print('- resultado_limpio.csv (formato CSV)')
print(f'- {excel_filename} (formato Excel con formato mejorado)')
print(f'Total de registros procesados: {len(df_final)}')

# Mostrar algunas estadísticas
print('\nResumen de datos:')
print(f'Número de índices únicos: {df_final["INDICE"].nunique()}')
print(f'Rango de fechas: {df_final["FECHA"].min()} a {df_final["FECHA"].max()}')
print(f'Índices incluidos: {", ".join(df_final["INDICE"].unique()[:5])}{"..." if df_final["INDICE"].nunique() > 5 else ""}')